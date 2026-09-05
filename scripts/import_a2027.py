#!/usr/bin/env python3
"""Import official BALI A-2027 price lists into the OYC web configurator.

Workbook contents are treated strictly as source data. The importer extracts
versions, the Excellence package, selectable options and preparation/delivery
items, then applies the established Polish terminology from the A-2026 catalog.
"""

from __future__ import annotations

import json
import re
import time
import unicodedata
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
CATALOG_PATH = ROOT / "data" / "models.json"
A27_DIR = Path("/Users/marekchorwacja/Library/Mobile Documents/com~apple~CloudDocs/Documents/bali/26/A27")
A26_DIR = Path("/Users/marekchorwacja/Library/Mobile Documents/com~apple~CloudDocs/Documents/bali/26/A26")


@dataclass(frozen=True)
class WorkbookConfig:
    model_id: str
    name: str
    filename: str
    sheet: str
    description_col: int
    price_col: int
    amount_col: int
    old_filename: str | None = None
    old_sheet: str | None = None
    old_description_col: int | None = None
    old_price_col: int | None = None
    tagline: str = ""


CONFIGS = [
    WorkbookConfig("bali-catsmart", "BALI CATSMART", "Tarif CATSMART - A 2027.xlsx", "Catsmart EN", 3, 4, 5, "Tarif CATSMART - A 2026.xlsx", "Catsmart EN", 3, 4, "BALI CATSMART wyposażony w 2 silniki Nanni 21 KM"),
    WorkbookConfig("bali-catspace", "BALI CATSPACE", "Tarif Bali CS A 2027.xlsx", "CS EN", 3, 4, 5, "Tarif Bali CS A 2026.xlsx", "CS EN", 3, 4, "BALI CATSPACE wyposażony w 2 silniki Nanni 19 KM"),
    WorkbookConfig("bali-4-2", "BALI 4.2", "Tarif Bali 4.2 A 2027.xlsx", "4.2 EN", 3, 4, 5, None, None, None, None, "BALI 4.2 wyposażony w 2 silniki Nanni 30 KM"),
    WorkbookConfig("bali-4-3", "BALI 4.3", "Tarif Bali 4.3 A 2027.xlsx", "4.3 EN", 2, 3, 4, "Tarif Bali 4.4 A 2026.xlsx", "4.4 EN", 3, 4, "BALI 4.3 wyposażony w 2 silniki 50 KM"),
    WorkbookConfig("bali-4-6", "BALI 4.6", "Tarif Bali 4.6 A 2027.xlsx", "4.6 EN", 3, 4, 5, "Tarif Bali 4.6 A 2026.xlsx", "4.6 EN", 3, 4, "BALI 4.6 wyposażony w 2 silniki Yanmar 45 KM"),
    WorkbookConfig("bali-5-2", "BALI 5.2", "Tarif Bali 5.2 A2027.xlsx", "ENGLISH", 2, 3, 4, "Tarif Bali 5.2 A2026 indice B.xlsx", "ENGLISH", 2, 3, "BALI 5.2 wyposażony w 2 silniki Yanmar 57 KM"),
    WorkbookConfig("bali-5-8", "BALI 5.8", "Tarif Bali 5.8 A 2027.xlsx", "5.8 EN", 3, 4, 5, "Tarif Bali 5.8 A 2026.xlsx", "5.8 EN", 3, 4, "BALI 5.8 wyposażony w 2 silniki Nanni 80 KM"),
    WorkbookConfig("bali-7-0", "BALI 7.0", "Tarif Bali 7.0 A2027.xlsx", "7.0 EN", 2, 3, 4, None, None, None, None, "BALI 7.0 wyposażony w 2 silniki Yanmar 195 KM"),
]


CATEGORY_TRANSLATIONS = {
    "rigging - sails": "Olinowanie i żagle",
    "mechanical - safety equipment": "Mechanika i wyposażenie bezpieczeństwa",
    "mechanics - safety equipment": "Mechanika i wyposażenie bezpieczeństwa",
    "comfort": "Komfort",
    "interior design": "Zabudowa wnętrza",
    "upholstery colors": "Kolor tapicerki",
    "upholstery colour": "Kolor tapicerki",
    "exterior design": "Wyposażenie zewnętrzne",
    "electronics - hi-fi": "Elektronika i audio",
    "electronics - hifi": "Elektronika i audio",
    "rigging- sails": "Olinowanie i żagle",
    "mecanics - safety equipment": "Mechanika i wyposażenie bezpieczeństwa",
    "mechanical equipment - safety equipment": "Mechanika i wyposażenie bezpieczeństwa",
    "engines and safety equipment": "Silniki i wyposażenie bezpieczeństwa",
    "interior setup": "Zabudowa wnętrza",
    "interior layout": "Zabudowa wnętrza",
    "interior fittings": "Zabudowa wnętrza",
    "upholstery color": "Kolor tapicerki",
    "upholstery colours": "Kolor tapicerki",
    "exterior setup": "Wyposażenie zewnętrzne",
    "exterior fittings": "Wyposażenie zewnętrzne",
    "preparation - delivery": "Przygotowanie i dostawa",
    "ambiance": "Stylistyka",
    "sailing and performance": "Żagle i osiągi",
    "propulsion and power": "Napęd i energia",
    "lifestyle": "Komfort i styl życia",
    "customised interior layout": "Indywidualna zabudowa wnętrza",
    "elegance": "Pakiet Elegance",
    "customised exterior layout": "Indywidualne wyposażenie zewnętrzne",
    "on-board technology": "Technologia pokładowa",
    "preparation - handover": "Przygotowanie i przekazanie",
}

TRANSLATION_CACHE = ROOT / "data" / "translations_a2027.json"


def text(value: Any) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def norm(value: Any) -> str:
    value = unicodedata.normalize("NFKD", text(value)).encode("ascii", "ignore").decode().lower()
    value = value.replace("portside", "port").replace("starboard side", "starboard")
    value = value.replace("heads compartments", "heads").replace("toilets", "heads")
    value = re.sub(r"\b(bali|option|total|price|pack|specifications?)\b", " ", value)
    value = re.sub(r"[^a-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip()


def is_price(value: Any) -> bool:
    return isinstance(value, (int, float)) and not isinstance(value, bool)


def demand(value: Any) -> bool:
    return "demand" in text(value).lower() or "request" in text(value).lower()


def locate(ws, needle: str, column: int) -> int:
    needle = norm(needle)
    for row in range(1, ws.max_row + 1):
        if needle in norm(ws.cell(row, column).value):
            return row
    raise ValueError(f"Nie znaleziono {needle!r} w arkuszu {ws.title}")


def locate_any(ws, needles: tuple[str, ...]) -> int:
    normalized = tuple(norm(item) for item in needles)
    for row in range(1, ws.max_row + 1):
        joined = " ".join(norm(ws.cell(row, col).value) for col in range(1, min(ws.max_column, 6) + 1))
        if any(needle in joined for needle in normalized):
            return row
    raise ValueError(f"Nie znaleziono {needles!r} w arkuszu {ws.title}")


def extract(ws, cfg: WorkbookConfig) -> dict[str, Any]:
    dcol, pcol, acol = cfg.description_col, cfg.price_col, cfg.amount_col
    is_bali_70 = cfg.model_id == "bali-7-0"
    excellence_header = locate_any(ws, ("customisation",)) if is_bali_70 else locate(ws, "excellence pack specifications", dcol)
    options_header = excellence_header if is_bali_70 else locate_any(ws, ("options",))
    delivery_header = locate_any(ws, ("preparation delivery", "preparation handover", "commissioning handover", "commissioning handing over"))

    versions = []
    for row in range(1, excellence_header):
        description, price = text(ws.cell(row, dcol).value), ws.cell(row, pcol).value
        if is_price(price) and "cabin" in description.lower():
            versions.append({"row": row, "en": description, "price": int(price)})

    included = []
    excellence_price = None
    standard_end = delivery_header if is_bali_70 else options_header
    for row in range(excellence_header + 1, standard_end):
        description, price = text(ws.cell(row, dcol).value), ws.cell(row, pcol).value
        if not description:
            continue
        if text(price).lower() == "x":
            included.append({"row": row, "en": description})
        elif is_price(price) and "excellence" in description.lower():
            excellence_price = int(price)
    if is_bali_70:
        excellence_price = 0
    elif excellence_price is None:
        raise ValueError(f"Brak ceny pakietu Excellence: {cfg.name}")

    options = []
    category = "Opcje"
    pending_note = ""
    for row in range(options_header + 1, delivery_header):
        marker = text(ws.cell(row, 1).value)
        description = text(ws.cell(row, dcol).value)
        price = ws.cell(row, pcol).value
        amount = ws.cell(row, acol).value
        if marker == "#" and description:
            translated_category = CATEGORY_TRANSLATIONS.get(description.lower().strip())
            if translated_category:
                category, pending_note = translated_category, ""
            else:
                pending_note = description
            continue
        if not description or not (is_price(price) or demand(price)):
            continue
        options.append({
            "row": row,
            "en": description,
            "price": int(price) if is_price(price) else None,
            "priceOnRequest": demand(price),
            "category": category,
            "note_en": pending_note,
        })

    delivery = []
    for row in range(delivery_header + 1, ws.max_row + 1):
        description = text(ws.cell(row, dcol).value)
        price = ws.cell(row, pcol).value
        if not description or not (is_price(price) or demand(price)):
            continue
        delivery.append({
            "row": row,
            "en": description,
            "price": int(price) if is_price(price) else None,
            "priceOnRequest": demand(price),
            "category": "Przygotowanie i dostawa",
            "note_en": "",
        })
    return {"versions": versions, "included": included, "excellence_price": excellence_price, "options": options, "delivery": delivery}


def google_translate(source: str) -> str:
    query = urllib.parse.urlencode({"client": "gtx", "sl": "en", "tl": "pl", "dt": "t", "q": source})
    url = f"https://translate.googleapis.com/translate_a/single?{query}"
    last_error: Exception | None = None
    for attempt in range(4):
        try:
            with urllib.request.urlopen(url, timeout=30) as response:
                payload = json.loads(response.read().decode("utf-8"))
            return "".join(part[0] for part in payload[0] if part and part[0]).strip()
        except Exception as exc:  # pragma: no cover - network retry
            last_error = exc
            time.sleep(1.5 * (attempt + 1))
    raise RuntimeError(f"Nie udało się przetłumaczyć: {source}") from last_error


def polish_cleanup(value: str) -> str:
    replacements = {
        "komora głowic": "toaleta",
        "przedziały głowic": "toalety",
        "przedziały głów": "toalety",
        "przedziały na toalety": "toalety",
        "przedziałów na toalety": "toalet",
        "przedziałów głowic": "toalet",
        "głowice": "toalety",
        "głowy": "toalety",
        "lewa burta": "lewa burta",
        "prawa burta": "prawa burta",
        "łódź przetargowa": "ponton",
        "przetarg": "ponton",
        "arkusz Solent": "szot Solenta",
        "arkusze": "szoty",
        "prześcieradła": "szoty",
        "pokrowiec na leniwą torbę": "Lazy Bag",
        "leniwa torba": "Lazy Bag",
        "leniwe podnośniki": "Lazy Jacks",
        "Kod 0": "Code 0",
        "kod 0": "Code 0",
        "pokój dzienny": "salon",
        "salonie": "salonie",
        "kabina właściciela": "kabina armatorska",
        "apartament właściciela": "apartament armatorski",
        "apartament właścicielski": "apartament armatorski",
        "APARTAMENT OWNER": "APARTAMENT ARMATORSKI",
        "OWNER’S SUITE": "ARMATORSKA",
        "sterówce": "stanowisku sternika",
        "do portu": "na lewej burcie",
        "port VIP": "VIP na lewej burcie",
        "VIP porty": "VIP na lewej burcie",
        "1 port": "1 lewa burta",
        "2 porty": "2 lewe burty",
        "3 porty": "3 lewe burty",
        "salon główny i prywatny": "apartament armatorski i prywatny salon",
        "Master i prywatny salon": "apartament armatorski i prywatny salon",
        "Master": "apartament armatorski",
        "platformy cukrowe": "rufa",
        "łyżki cukru": "rufa",
        "cukierki": "rufy",
        "VAT bez VAT": "netto",
        "Cena bez podatku": "Cena netto",
        "Filtr do oczyszczania świeżej wody": "Filtr oczyszczający wodę słodką",
        "Kontroler baterii": "Monitor akumulatorów",
        "Monitor baterii": "Monitor akumulatorów",
        "baterie serwisowe": "akumulatory serwisowe",
        "baterii serwisowej": "akumulatorów serwisowych",
        "ekranem dźwiękowym": "obudową wygłuszającą",
        "osłoną dźwiękową": "obudową wygłuszającą",
        "70 ml": "70 m",
        "80 ml": "80 m",
        "Raymarine Electronic Pack": "Pakiet elektroniki Raymarine",
        "Hifi Radio Fusion": "System audio Fusion",
        "Pakiet Total ELEGANCE": "Pakiet Elegance",
        "Pakiet Totalnej Elegancji": "Pakiet Elegance",
        "Łącznie za Pakiet Elegancji": "Pakiet Elegance",
        "Pakiet Elegancji Razem": "Pakiet Elegance",
        "Łączny pakiet Elegancji": "Pakiet Elegance",
        "Razem pakiet ELEGANCJA": "Pakiet Elegance",
        "kwadratowym dachem": "kwadratowym wierzchołkiem",
        "duch dziobowy": "bukszpryt",
        "wejście na maszt": "postawienie masztu",
        "skarpetki na błotniki": "pokrowce na odbijacze",
        "skarpetek na błotniki": "pokrowców na odbijacze",
        "Owijka +": "Oklejenie ochronne +",
        "odbiór i leczenie": "odbiór i obsługa",
        "Pakiet Gotowy do pracy": "Pakiet Ready to Go",
        "Gotowy do użycia pakiet": "Pakiet Ready to Go",
        "Gotowy pakiet": "Pakiet Ready to Go",
    }
    result = text(value)
    for source, target in replacements.items():
        result = result.replace(source, target)
    result = re.sub(r"\b([23456])-kabinowa wersja\b", r"Wersja \1-kabinowa", result, flags=re.I)
    result = re.sub(r"\b([23456]) kabinowa wersja\b", r"Wersja \1-kabinowa", result, flags=re.I)
    result = re.sub(r"Wersja z ([23456]) kabinami", r"Wersja \1-kabinowa", result, flags=re.I)
    result = re.sub(r"Wersja ([23456]) kabinowa", r"Wersja \1-kabinowa", result, flags=re.I)
    return result


def translate_all(strings: set[str]) -> dict[str, str]:
    cache = json.loads(TRANSLATION_CACHE.read_text()) if TRANSLATION_CACHE.exists() else {}
    missing = sorted(item for item in strings if item and item not in cache)
    if missing:
        print(f"Tłumaczenie nowych opisów: {len(missing)}")
        with ThreadPoolExecutor(max_workers=8) as pool:
            future_map = {pool.submit(google_translate, item): item for item in missing}
            for index, future in enumerate(as_completed(future_map), 1):
                source = future_map[future]
                cache[source] = polish_cleanup(future.result())
                if index % 100 == 0:
                    print(f"  {index}/{len(missing)}")
        TRANSLATION_CACHE.write_text(json.dumps(cache, ensure_ascii=False, indent=2) + "\n")
    return {source: polish_cleanup(cache[source]) for source in strings if source}


def model_from_parsed(cfg: WorkbookConfig, parsed: dict[str, Any], translations: dict[str, str]) -> dict[str, Any]:
    def translated(source: str) -> str:
        return translations.get(source, source)

    def option_from(item: dict[str, Any]) -> dict[str, Any]:
        note = translated(item["note_en"]) if item["note_en"] else ("Cena na zapytanie" if item["priceOnRequest"] else "")
        return {
            "id": f"{cfg.model_id}-{item['row']}",
            "sourceRow": item["row"],
            "description": translated(item["en"]),
            "price": item["price"],
            "priceOnRequest": item["priceOnRequest"],
            "defaultQuantity": 0,
            "category": item["category"],
            "note": note,
        }

    excellence_name = "Wyposażenie standardowe" if cfg.model_id == "bali-7-0" else "Pakiet Excellence"
    return {
        "id": cfg.model_id,
        "name": cfg.name,
        "tagline": cfg.tagline,
        "versions": [
            {
                "id": f"{cfg.model_id}-v{index}",
                "name": translated(item["en"]),
                "basePrice": item["price"],
                "standardEngines": cfg.tagline.split("wyposażony w ", 1)[-1],
            }
            for index, item in enumerate(parsed["versions"], 1)
        ],
        "excellencePackage": {
            "name": excellence_name,
            "price": parsed["excellence_price"],
            "included": [{"sourceRow": item["row"], "description": translated(item["en"])} for item in parsed["included"]],
        },
        "options": [option_from(item) for item in parsed["options"]],
        "delivery": [option_from(item) for item in parsed["delivery"]],
        "sourceSheet": cfg.sheet,
    }


def main() -> None:
    parsed_models: list[tuple[WorkbookConfig, dict[str, Any]]] = []
    strings: set[str] = set()
    for cfg in CONFIGS:
        wb = openpyxl.load_workbook(A27_DIR / cfg.filename, read_only=True, data_only=False)
        parsed = extract(wb[cfg.sheet], cfg)
        print(cfg.name, {k: len(v) if isinstance(v, list) else v for k, v in parsed.items()})
        parsed_models.append((cfg, parsed))
        for group in ("versions", "included", "options", "delivery"):
            for item in parsed[group]:
                strings.add(item["en"])
                if item.get("note_en"):
                    strings.add(item["note_en"])

    translations = translate_all(strings)
    output = {
        "source": "Oficjalne cenniki BALI A-2027",
        "priceList": "A-2027",
        "currency": "EUR",
        "netPrices": True,
        "models": [model_from_parsed(cfg, parsed, translations) for cfg, parsed in parsed_models],
    }
    CATALOG_PATH.write_text(json.dumps(output, ensure_ascii=False, indent=2) + "\n")
    print(f"Zapisano {len(output['models'])} modeli w {CATALOG_PATH}")


if __name__ == "__main__":
    main()
